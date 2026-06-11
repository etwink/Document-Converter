#!/usr/bin/env python3

import json
import re
import shutil
import sys
from pathlib import Path

import pandas as pd

ACCEPTED_FILE_TYPES = [
    'c', 'cpp', 'doc', 'docx', 'html', 'java', 'json', 
    'md', 'pdf', 'php', 'pptx', 'py', 'rb', 'tex', 'txt', 
    'css', 'js', 'sh', 'ts'
]

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xlsb"}
CSV_EXTENSIONS = {".csv"}
COBOL_EXTENSIONS = {".cbl", ".cob", ".cobol", ".cpy", ".cobc", ".mps", ".src", ".ct1", ".jcv", ".prv"}
ACCEPTED_EXTENSIONS = {f".{ext}" for ext in ACCEPTED_FILE_TYPES}

def sanitize_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(name)).strip(". ")

def _check_is_empty_formula_file(data: str) -> bool:
    lines = data.splitlines()
    # lines[0] is the header row — skip it, only check data rows
    return all(line.replace(',', '').strip() == '' for line in lines[1:])

def _xlsb_cell_formula(cell):
    try:
        return f"={cell.f}" if cell.f else cell.v
    except AttributeError:
        return None


def _extract_vba_macros(input_path: Path, output_dir: Path) -> list[Path]:
    try:
        from oletools.olevba import VBA_Parser
    except ImportError:
        print(f"[WARN]  oletools not installed — skipping macro extraction for {input_path.name}", file=sys.stderr)
        return []

    vba_parser = VBA_Parser(str(input_path))
    if not vba_parser.detect_vba_macros():
        return []

    modules = []
    for (_, _, vba_filename, vba_code) in vba_parser.extract_macros():
        modules.append(f"' === {vba_filename} ===\n{vba_code}")

    if not modules:
        return []

    out = output_dir / f"{input_path.stem}_macro.txt"
    out.write_text('\n\n'.join(modules), encoding='utf-8')
    return [out]


def convert_excel(input_path: Path, output_dir: Path) -> list[Path]:
    output_files = []

    if input_path.suffix.lower() == ".xlsb":
        import pyxlsb
        with pyxlsb.open_workbook(str(input_path)) as wb:
            for sheet_name in wb.sheets:
                with wb.get_sheet(sheet_name) as sheet:
                    all_rows = list(sheet.rows())
                if not all_rows:
                    continue

                data_rows = [[cell.v for cell in row] for row in all_rows]
                formula_rows = [
                    [_xlsb_cell_formula(cell) for cell in row]
                    for row in all_rows
                ]

                safe_name = sanitize_filename(sheet_name)
                base = f"{input_path.stem}_{safe_name}"

                out_data = output_dir / f"{base}_data.txt"
                out_formula = output_dir / f"{base}_formula.txt"

                pd.DataFrame(data_rows[1:], columns=data_rows[0]).to_csv(out_data, index=False)
                if _check_is_empty_formula_file(pd.DataFrame(formula_rows[1:], columns=formula_rows[0]).to_csv(index=False)):
                    print(f"[SKIP]  {out_formula.name} (no formulas found)")
                    output_files.extend([out_data])
                else:
                    pd.DataFrame(formula_rows[1:], columns=formula_rows[0]).to_csv(out_formula, index=False)
                    output_files.extend([out_data, out_formula])
    else:
        import openpyxl
        wb_data = openpyxl.load_workbook(input_path, data_only=True)
        wb_formula = openpyxl.load_workbook(input_path, data_only=False)

        for sheet_name in wb_data.sheetnames:
            ws_data = wb_data[sheet_name]
            ws_formula = wb_formula[sheet_name]

            safe_name = sanitize_filename(sheet_name)
            base = f"{input_path.stem}_{safe_name}"

            data_rows = [[cell.value for cell in row] for row in ws_data.iter_rows()]
            formula_rows = [[cell.value for cell in row] for row in ws_formula.iter_rows()]

            if data_rows:
                out_data = output_dir / f"{base}_data.txt"
                pd.DataFrame(data_rows[1:], columns=data_rows[0]).to_csv(out_data, index=False)
                output_files.append(out_data)

            if formula_rows:
                out_formula = output_dir / f"{base}_formula.txt"
                if _check_is_empty_formula_file(pd.DataFrame(formula_rows[1:], columns=formula_rows[0]).to_csv(index=False)):
                    print(f"[SKIP]  {out_formula.name} (no formulas found)")
                else:
                    pd.DataFrame(formula_rows[1:], columns=formula_rows[0]).to_csv(out_formula, index=False)
                    output_files.append(out_formula)

        if input_path.suffix.lower() == ".xlsm":
            output_files.extend(_extract_vba_macros(input_path, output_dir))

    return output_files


def convert_csv(input_path: Path, output_dir: Path) -> list[Path]:
    df = pd.read_csv(input_path)
    out = output_dir / f"{input_path.stem}.txt"
    df.to_csv(out, index=False)
    return [out]


def convert_cobol(input_path: Path, output_dir: Path) -> list[Path]:
    contents = input_path.read_text(encoding="utf-8", errors="replace")
    data = {
        "File Name": f"{input_path.stem}{input_path.suffix}",
        "File Contents": contents,
    }
    out = output_dir / f"{input_path.stem}.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, indent="\t", ensure_ascii=False)
    return [out]


def main(source_dir: str) -> None:
    source_path = Path(source_dir)
    if not source_path.is_dir():
        print(f"Error: '{source_dir}' is not a valid directory.", file=sys.stderr)
        sys.exit(1)

    output_dir = source_path.parent / f"{source_path.name}_converted"
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {output_dir}\n")

    converted: list[Path] = []
    errors: list[tuple[Path, Exception]] = []

    for file_path in sorted(source_path.rglob("*")):
        if not file_path.is_file():
            continue

        ext = file_path.suffix.lower()

        try:
            if ext in EXCEL_EXTENSIONS:
                outputs = convert_excel(file_path, output_dir)
            elif ext in CSV_EXTENSIONS:
                outputs = convert_csv(file_path, output_dir)
            elif ext in COBOL_EXTENSIONS:
                outputs = convert_cobol(file_path, output_dir)
            elif ext in ACCEPTED_EXTENSIONS:
                out = output_dir / file_path.name
                shutil.copy2(file_path, out)
                outputs = [out]
            else:
                print(f"[SKIP]  {file_path.name}")
                continue

            for out in outputs:
                print(f"[OK]    {file_path.name} -> {out.name}")
            converted.extend(outputs)

        except Exception as exc:
            print(f"[ERROR] {file_path.name}: {exc}", file=sys.stderr)
            errors.append((file_path, exc))

    print(f"\n{len(converted)} file(s) written to '{output_dir}'", end="")
    if errors:
        print(f", {len(errors)} error(s).")
    else:
        print(".")


if __name__ == "__main__":
    dir_name = r"C:\path\to\your\folder"  # Edit this to your source directory
    main(dir_name)
